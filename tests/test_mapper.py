"""Unit tests for ContactMapper service."""

import json
import pytest
from io import BytesIO
from pathlib import Path
from openpyxl import Workbook

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from models.schemas import Recipient, Contact
from services.contact_mapper import ContactMapper


def _create_contacts_workbook() -> bytes:
    """Create a sample contacts Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Tutor")
    ws.cell(row=1, column=2, value="Email")
    ws.cell(row=1, column=3, value="CC")

    contacts = [
        ("Oscar Herranz", "oscar.herranz@test.com", "cc1@test.com"),
        ("Juan Berral", "juan.berral@test.com", None),
        ("Maria Gregorio", "maria.gregorio@test.com", "cc2@test.com"),
    ]
    for i, (name, email, cc) in enumerate(contacts, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=email)
        if cc:
            ws.cell(row=i, column=3, value=cc)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


class TestContactMapper:
    """Tests for ContactMapper."""

    def test_load_contacts_from_bytes(self):
        """Test loading contacts from file content."""
        content = _create_contacts_workbook()
        mapper = ContactMapper()
        contacts = mapper.load_contacts(file_content=content, use_stored=False)

        assert len(contacts) == 3

    def test_map_exact_match(self):
        """Test exact name matching."""
        content = _create_contacts_workbook()
        mapper = ContactMapper()
        mapper.load_contacts(file_content=content, use_stored=False)

        recipients = [
            Recipient(codigo="Oscar Herranz", nombre="Oscar Herranz"),
            Recipient(codigo="Juan Berral", nombre="Juan Berral"),
        ]
        mappings = mapper.map_contacts(recipients)

        assert len(mappings) == 2
        assert mappings[0].email_found is True
        assert mappings[1].email_found is True
        assert mappings[0].contact.email == "oscar.herranz@test.com"

    def test_map_fuzzy_match_accents(self):
        """Test matching with accent differences."""
        content = _create_contacts_workbook()
        mapper = ContactMapper()
        mapper.load_contacts(file_content=content, use_stored=False)

        recipients = [
            Recipient(codigo="Óscar Herranz", nombre="Óscar Herranz"),
        ]
        mappings = mapper.map_contacts(recipients)

        assert mappings[0].email_found is True

    def test_unmatched_recipient(self):
        """Test unmatched recipient returns no contact."""
        content = _create_contacts_workbook()
        mapper = ContactMapper()
        mapper.load_contacts(file_content=content, use_stored=False)

        recipients = [
            Recipient(codigo="Ana Pidal", nombre="Ana Pidal"),
        ]
        mappings = mapper.map_contacts(recipients)

        assert mappings[0].email_found is False
        assert mappings[0].contact is None

    def test_persist_and_load(self, tmp_path):
        """Test contact persistence to JSON."""
        store_path = tmp_path / "contacts_store.json"

        # Monkey-patch settings
        import config
        original_path = config.settings.contacts_store_path
        config.settings.contacts_store_path = str(store_path)

        try:
            content = _create_contacts_workbook()
            mapper = ContactMapper()
            mapper.load_contacts(file_content=content, use_stored=False)

            # Verify file was created
            assert store_path.exists()

            # Load again from store
            mapper2 = ContactMapper()
            contacts = mapper2.load_contacts(use_stored=True)
            assert len(contacts) == 3
        finally:
            config.settings.contacts_store_path = original_path

    def test_delete_contacts_wrong_password(self, tmp_path):
        """Test deletion fails with wrong password."""
        store_path = tmp_path / "contacts_store.json"

        import config
        original_path = config.settings.contacts_store_path
        config.settings.contacts_store_path = str(store_path)

        try:
            content = _create_contacts_workbook()
            mapper = ContactMapper()
            mapper.load_contacts(file_content=content, use_stored=False)

            result = mapper.delete_stored_contacts("wrong_password")
            assert result is False
            assert store_path.exists()
        finally:
            config.settings.contacts_store_path = original_path

    def test_delete_contacts_correct_password(self, tmp_path):
        """Test deletion succeeds with correct password."""
        store_path = tmp_path / "contacts_store.json"

        import config
        original_path = config.settings.contacts_store_path
        original_pw = config.settings.contacts_delete_password
        config.settings.contacts_store_path = str(store_path)

        try:
            content = _create_contacts_workbook()
            mapper = ContactMapper()
            mapper.load_contacts(file_content=content, use_stored=False)

            result = mapper.delete_stored_contacts(config.settings.contacts_delete_password)
            assert result is True
            assert not store_path.exists()
        finally:
            config.settings.contacts_store_path = original_path

    def test_cc_email_extracted(self):
        """Test CC email is correctly extracted."""
        content = _create_contacts_workbook()
        mapper = ContactMapper()
        mapper.load_contacts(file_content=content, use_stored=False)

        recipients = [Recipient(codigo="Oscar Herranz", nombre="Oscar Herranz")]
        mappings = mapper.map_contacts(recipients)

        assert mappings[0].contact.email_cc == "cc1@test.com"
