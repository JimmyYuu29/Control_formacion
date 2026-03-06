"""Unit tests for ExcelParser service."""

import pytest
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from services.excel_parser import ExcelParser


def _create_sample_workbook() -> bytes:
    """Create a minimal test workbook with multi-row headers and Tutor column."""
    wb = Workbook()
    ws = wb.active

    # Row 1: config/threshold values (mostly empty)
    ws.cell(row=1, column=1, value=None)

    # Row 2: categories (merged cells simulated)
    ws.cell(row=2, column=3, value="Datos Personales")
    ws.cell(row=2, column=7, value="Nota Formacion")

    # Row 3: subcategories
    ws.cell(row=3, column=3, value="Profesional")
    ws.cell(row=3, column=4, value="DNI")
    ws.cell(row=3, column=5, value="MAIL")
    ws.cell(row=3, column=6, value="Tutor")
    ws.cell(row=3, column=7, value="Curso 1")
    ws.cell(row=3, column=8, value="Curso 2")
    ws.cell(row=3, column=9, value="Total")

    # Row 4: empty separator
    # Row 5+: data
    data = [
        (5, "Maria Grande", "12345678A", "maria@test.com", "Oscar Herranz", 2, 2, 4),
        (6, "Marcos Rios", "23456789B", "marcos@test.com", "Juan Berral", 2, 1, 3),
        (7, "Miguel Antelo", "34567890C", "miguel@test.com", "Juan Berral", 1, 2, 3),
        (8, "Angela Montilla", "45678901D", "angela@test.com", "Maria Gregorio", 2, 2, 4),
        (9, "Pedro Lopez", "56789012E", "pedro@test.com", "Oscar Herranz", 1, 1, 2),
    ]
    for row_idx, nombre, dni, mail, tutor, c1, c2, total in data:
        ws.cell(row=row_idx, column=3, value=nombre)
        ws.cell(row=row_idx, column=4, value=dni)
        ws.cell(row=row_idx, column=5, value=mail)
        ws.cell(row=row_idx, column=6, value=tutor)
        ws.cell(row=row_idx, column=7, value=c1)
        ws.cell(row=row_idx, column=8, value=c2)
        ws.cell(row=row_idx, column=9, value=total)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


class TestExcelParser:
    """Tests for ExcelParser."""

    def test_parse_valid_file(self):
        """Test parsing a valid file detects all tutors."""
        content = _create_sample_workbook()
        parser = ExcelParser(mode="tutor")
        doc = parser.parse(content, "test.xlsx")

        assert doc.filename == "test.xlsx"
        assert len(doc.blocks) == 3  # Oscar, Juan, Maria
        assert doc.tutor_column is not None

    def test_groups_correct_counts(self):
        """Test each tutor group has correct number of entries."""
        content = _create_sample_workbook()
        parser = ExcelParser()
        doc = parser.parse(content, "test.xlsx")

        counts = {b.recipient.nombre: len(b.entries) for b in doc.blocks}
        assert counts["Oscar Herranz"] == 2
        assert counts["Juan Berral"] == 2
        assert counts["Maria Gregorio"] == 1

    def test_column_info_extracted(self):
        """Test column metadata is extracted."""
        content = _create_sample_workbook()
        parser = ExcelParser()
        doc = parser.parse(content, "test.xlsx")

        assert len(doc.all_columns) > 0
        letters = [c.letter for c in doc.all_columns]
        assert "F" in letters  # Tutor column

    def test_data_entries_have_values(self):
        """Test data entries contain raw cell values."""
        content = _create_sample_workbook()
        parser = ExcelParser()
        doc = parser.parse(content, "test.xlsx")

        block = next(b for b in doc.blocks if b.recipient.nombre == "Juan Berral")
        assert len(block.entries) == 2
        # Check first entry has raw_cells
        entry = block.entries[0]
        assert "C" in entry.raw_cells or "D" in entry.raw_cells

    def test_no_tutor_column_raises(self):
        """Test error when Tutor column is missing."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=3, column=1, value="Nombre")
        ws.cell(row=3, column=2, value="DNI")
        ws.cell(row=5, column=1, value="Test")
        ws.cell(row=5, column=2, value="123")

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        parser = ExcelParser()
        with pytest.raises(ValueError, match="Tutor"):
            parser.parse(buf.read(), "bad.xlsx")

    def test_empty_rows_skipped(self):
        """Test that empty separator rows are skipped."""
        content = _create_sample_workbook()
        parser = ExcelParser()
        doc = parser.parse(content, "test.xlsx")

        # All entries should have source_row >= 5 (after headers + separator)
        for block in doc.blocks:
            for entry in block.entries:
                assert entry.source_row >= 5

    def test_header_rows_set(self):
        """Test header_rows is correctly set."""
        content = _create_sample_workbook()
        parser = ExcelParser()
        doc = parser.parse(content, "test.xlsx")

        assert doc.header_rows == [1, 2, 3]

    def test_default_columns_include_all(self):
        """Test default columns include all detected columns."""
        content = _create_sample_workbook()
        parser = ExcelParser()
        doc = parser.parse(content, "test.xlsx")

        assert len(doc.default_columns) == len(doc.all_columns)
