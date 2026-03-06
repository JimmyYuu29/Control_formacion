"""Unit tests for ExcelGenerator service."""

import pytest
from io import BytesIO
from openpyxl import Workbook, load_workbook

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from services.excel_parser import ExcelParser
from services.excel_generator import ExcelGenerator


def _create_sample_workbook() -> bytes:
    """Create a sample workbook for testing."""
    wb = Workbook()
    ws = wb.active

    # Row 2: categories
    ws.cell(row=2, column=3, value="Datos")
    ws.cell(row=2, column=7, value="Notas")

    # Row 3: subcategories
    ws.cell(row=3, column=3, value="Profesional")
    ws.cell(row=3, column=4, value="DNI")
    ws.cell(row=3, column=5, value="MAIL")
    ws.cell(row=3, column=6, value="Tutor")
    ws.cell(row=3, column=7, value="Curso 1")
    ws.cell(row=3, column=8, value="Total")

    # Data
    data = [
        (5, "Maria Grande", "12345678A", "maria@t.com", "Oscar Herranz", 8, 8),
        (6, "Marcos Rios", "23456789B", "marcos@t.com", "Juan Berral", 7, 7),
        (7, "Miguel Antelo", "34567890C", "miguel@t.com", "Juan Berral", 6, 6),
    ]
    for row_idx, nombre, dni, mail, tutor, c1, total in data:
        ws.cell(row=row_idx, column=3, value=nombre)
        ws.cell(row=row_idx, column=4, value=dni)
        ws.cell(row=row_idx, column=5, value=mail)
        ws.cell(row=row_idx, column=6, value=tutor)
        ws.cell(row=row_idx, column=7, value=c1)
        ws.cell(row=row_idx, column=8, value=total)

    # Set column widths
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


class TestExcelGenerator:
    """Tests for ExcelGenerator."""

    def test_generate_files_count(self):
        """Test correct number of files generated."""
        content = _create_sample_workbook()
        parser = ExcelParser()
        doc = parser.parse(content, "test.xlsx")

        gen = ExcelGenerator()
        files = gen.generate_files(doc, content, doc.default_columns, "test.xlsx")

        # Should be 2 tutors: Oscar (1 entry), Juan (2 entries)
        assert len(files) == 2

    def test_generated_file_has_headers(self):
        """Test generated files contain header rows."""
        content = _create_sample_workbook()
        parser = ExcelParser()
        doc = parser.parse(content, "test.xlsx")

        gen = ExcelGenerator()
        files = gen.generate_files(doc, content, doc.default_columns, "test.xlsx")

        filename, file_content = files[0]
        wb = load_workbook(BytesIO(file_content))
        ws = wb.active

        # Should have at least 3 rows (headers) + data rows
        assert ws.max_row >= 4
        wb.close()

    def test_generated_filename_contains_tutor(self):
        """Test filenames contain tutor name."""
        content = _create_sample_workbook()
        parser = ExcelParser()
        doc = parser.parse(content, "test.xlsx")

        gen = ExcelGenerator()
        files = gen.generate_files(doc, content, doc.default_columns, "test.xlsx")

        filenames = [f[0] for f in files]
        assert any("Oscar" in fn for fn in filenames)
        assert any("Juan" in fn for fn in filenames)

    def test_column_selection(self):
        """Test only selected columns appear in output."""
        content = _create_sample_workbook()
        parser = ExcelParser()
        doc = parser.parse(content, "test.xlsx")

        # Select only C, F columns
        gen = ExcelGenerator()
        files = gen.generate_files(doc, content, ["C", "F"], "test.xlsx")

        filename, file_content = files[0]
        wb = load_workbook(BytesIO(file_content))
        ws = wb.active

        # Should have max 2 columns of data
        assert ws.max_column <= 2
        wb.close()

    def test_create_zip(self):
        """Test ZIP archive creation."""
        gen = ExcelGenerator()
        files = [("test1.xlsx", b"content1"), ("test2.xlsx", b"content2")]
        zip_content = gen.create_zip_archive(files)

        assert len(zip_content) > 0
        assert zip_content[:2] == b"PK"  # ZIP magic number

    def test_excluded_blocks_skipped(self):
        """Test excluded blocks are not generated."""
        content = _create_sample_workbook()
        parser = ExcelParser()
        doc = parser.parse(content, "test.xlsx")

        # Exclude first tutor
        doc.blocks[0].recipient.excluded = True

        gen = ExcelGenerator()
        files = gen.generate_files(doc, content, doc.default_columns, "test.xlsx")

        assert len(files) == 1  # Only 1 of 2 tutors
