"""Contact mapper service for matching Tutors with contact information."""

import json
import logging
import re
import unicodedata
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

from config import settings
from models.schemas import Recipient, Contact, ContactMapping

logger = logging.getLogger(__name__)


class ContactMapper:
    """Service for mapping Tutors to contact information.

    Supports persistent contact storage: contacts are accumulated across
    imports and stored in a JSON file. Deletion requires a password.
    """

    def __init__(self, contacts_file: Optional[str] = None):
        """Initialize the contact mapper."""
        self.contacts_file = contacts_file or settings.contacts_file_path
        self._contacts_cache: Dict[str, Contact] = {}
        self._loaded = False

    # ── Persistence ──────────────────────────────────────────────────

    def _get_store_path(self) -> Path:
        """Get the path to the contacts store JSON file."""
        return Path(settings.contacts_store_path)

    def _load_stored_contacts(self) -> Dict[str, Contact]:
        """Load persisted contacts from JSON file."""
        store_path = self._get_store_path()
        if not store_path.exists():
            return {}
        try:
            with open(store_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            contacts = {}
            for key, contact_dict in data.get("contacts", {}).items():
                contacts[key] = Contact(**contact_dict)
            return contacts
        except (json.JSONDecodeError, Exception):
            return {}

    def _save_contacts_to_store(self) -> None:
        """Save current contacts to JSON file."""
        store_path = self._get_store_path()
        store_path.parent.mkdir(parents=True, exist_ok=True)
        data = {
            "contacts": {
                key: contact.model_dump(mode="json")
                for key, contact in self._contacts_cache.items()
            },
            "last_updated": datetime.now().isoformat(),
        }
        with open(store_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)

    def delete_stored_contacts(self, password: str) -> bool:
        """Delete stored contacts if password is correct.

        Args:
            password: Password to verify

        Returns:
            True if deletion succeeded, False if wrong password
        """
        if password != settings.contacts_delete_password:
            return False
        store_path = self._get_store_path()
        if store_path.exists():
            store_path.unlink()
        self._contacts_cache.clear()
        self._loaded = False
        logger.info("Contactos almacenados eliminados")
        return True

    def get_stored_contacts_info(self) -> dict:
        """Return info about stored contacts."""
        store_path = self._get_store_path()
        if not store_path.exists():
            return {"has_stored": False, "count": 0, "last_updated": None}
        try:
            with open(store_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return {
                "has_stored": True,
                "count": len(data.get("contacts", {})),
                "last_updated": data.get("last_updated"),
            }
        except (json.JSONDecodeError, Exception):
            return {"has_stored": False, "count": 0, "last_updated": None}

    # ── Loading ──────────────────────────────────────────────────────

    def load_contacts(
        self,
        file_path: Optional[str] = None,
        file_content: Optional[bytes] = None,
        use_stored: bool = True,
    ) -> List[Contact]:
        """Load contacts with accumulation from persistent store.

        1. Load previously stored contacts
        2. Parse new contacts from Excel (if available)
        3. Merge (new overrides stored on key conflict)
        4. Save merged result

        Args:
            file_path: Path to contacts file
            file_content: File content as bytes
            use_stored: Whether to load stored contacts first

        Returns:
            List of loaded contacts
        """
        # Step 1: Load stored contacts
        stored = self._load_stored_contacts() if use_stored else {}

        # Step 2: Parse new contacts from Excel
        new_contacts = {}
        try:
            new_contacts = self._parse_contacts_from_excel(file_path, file_content)
        except FileNotFoundError:
            if not file_path and not file_content:
                pass  # No file — use only stored contacts
            else:
                raise

        # Step 3: Merge
        merged = {**stored}
        for key, contact in new_contacts.items():
            merged[key] = contact

        self._contacts_cache = merged
        self._loaded = True

        # Step 4: Persist
        if merged:
            self._save_contacts_to_store()

        logger.info("Contactos cargados: %d total", len(self._contacts_cache))
        return list(self._contacts_cache.values())

    def _parse_contacts_from_excel(
        self,
        file_path: Optional[str] = None,
        file_content: Optional[bytes] = None,
    ) -> Dict[str, Contact]:
        """Parse contacts from an Excel file.

        Expected columns: Tutor Name (A), Email (B), CC Email (C)
        """
        contacts: Dict[str, Contact] = {}

        if file_content:
            wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        elif file_path:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        else:
            path = Path(self.contacts_file)
            if not path.exists():
                raise FileNotFoundError(
                    f"Archivo de contactos no encontrado: {path}"
                )
            wb = openpyxl.load_workbook(str(path), data_only=True)

        ws = wb.active

        try:
            header_row, col_map = self._find_headers(ws)
            if not col_map:
                raise ValueError(
                    "No se encontraron las columnas requeridas en el archivo de contactos. "
                    "Se necesita al menos: nombre del tutor y email."
                )

            for row_idx in range(header_row + 1, ws.max_row + 1):
                contact = self._parse_contact_row(ws, row_idx, col_map)
                if contact:
                    key = self._normalize_name(contact.nombre)
                    contacts[key] = contact

            return contacts
        finally:
            wb.close()

    def _find_headers(self, ws) -> tuple:
        """Find header row and column mapping."""
        expected_cols = {
            "tutor": "nombre",
            "nombre": "nombre",
            "responsable": "nombre",
            "email": "email",
            "e-mail": "email",
            "correo": "email",
            "mail": "email",
            "cc": "email_cc",
            "copia": "email_cc",
        }

        for row_idx in range(1, min(6, ws.max_row + 1)):
            col_map: Dict[str, int] = {}

            for col_idx in range(1, min(10, ws.max_column + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is None:
                    continue
                cell_text = str(cell_value).lower().strip()

                for keyword, field_name in expected_cols.items():
                    if keyword in cell_text:
                        if field_name == "email_cc" and "email" in col_map:
                            col_map["email_cc"] = col_idx
                        elif field_name not in col_map:
                            col_map[field_name] = col_idx
                        break

            if "nombre" in col_map and "email" in col_map:
                return row_idx, col_map

        # Fallback: assume A=name, B=email, C=cc
        return 0, {"nombre": 1, "email": 2, "email_cc": 3}

    def _parse_contact_row(
        self, ws, row_idx: int, col_map: Dict[str, int]
    ) -> Optional[Contact]:
        """Parse a single contact row."""
        nombre_cell = ws.cell(row=row_idx, column=col_map["nombre"]).value
        if nombre_cell is None or (isinstance(nombre_cell, str) and not nombre_cell.strip()):
            return None
        nombre = str(nombre_cell).strip()

        email_cell = ws.cell(row=row_idx, column=col_map["email"]).value
        if email_cell is None:
            return None
        email = str(email_cell).strip()
        if not self._is_valid_email(email):
            return None

        email_cc = None
        if "email_cc" in col_map:
            cc_cell = ws.cell(row=row_idx, column=col_map["email_cc"]).value
            if cc_cell:
                cc_email = str(cc_cell).strip()
                if self._is_valid_email(cc_email):
                    email_cc = cc_email

        return Contact(
            codigo=nombre,
            nombre=nombre,
            email=email,
            email_cc=email_cc,
        )

    # ── Matching ─────────────────────────────────────────────────────

    def _normalize_name(self, name: str) -> str:
        """Normalize a name for matching (remove accents, uppercase, collapse spaces)."""
        normalized = name.strip().upper()
        normalized = unicodedata.normalize("NFKD", normalized)
        normalized = "".join(c for c in normalized if not unicodedata.combining(c))
        normalized = " ".join(normalized.split())
        return normalized

    def _is_valid_email(self, email: str) -> bool:
        """Check if email has valid format."""
        pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
        return bool(re.match(pattern, email))

    def map_contacts(
        self,
        recipients: List[Recipient],
        contacts: Optional[List[Contact]] = None,
    ) -> List[ContactMapping]:
        """Map recipients (Tutors) to contacts.

        Args:
            recipients: List of Tutor recipients from parsed document
            contacts: Optional list of contacts (uses cache if None)

        Returns:
            List of ContactMapping results
        """
        if not self._loaded:
            self.load_contacts()

        mappings: List[ContactMapping] = []
        for recipient in recipients:
            contact = self._find_contact_for(recipient)
            mappings.append(ContactMapping(
                recipient=recipient,
                contact=contact,
                email_found=contact is not None,
                excluded=recipient.excluded,
            ))

        return mappings

    def _find_contact_for(self, recipient: Recipient) -> Optional[Contact]:
        """Find the best matching contact for a recipient.

        Tries exact name match first, then fuzzy match.
        """
        normalized_name = self._normalize_name(recipient.nombre)

        # Exact match
        if normalized_name in self._contacts_cache:
            return self._contacts_cache[normalized_name]

        # Fuzzy match: check if one contains the other
        for key, contact in self._contacts_cache.items():
            contact_name = self._normalize_name(contact.nombre_completo)
            if contact_name == normalized_name:
                return contact
            if contact_name in normalized_name or normalized_name in contact_name:
                return contact

        # Token-based match: check if all tokens of one name appear in the other
        recipient_tokens = set(normalized_name.split())
        for key, contact in self._contacts_cache.items():
            contact_tokens = set(self._normalize_name(contact.nombre_completo).split())
            if recipient_tokens and contact_tokens:
                if recipient_tokens.issubset(contact_tokens) or contact_tokens.issubset(recipient_tokens):
                    return contact

        return None

    def persist_contacts(self, contacts: List[Contact]) -> None:
        """Save a list of contacts to the persistent store."""
        for contact in contacts:
            key = self._normalize_name(contact.nombre)
            self._contacts_cache[key] = contact
        self._save_contacts_to_store()

    def get_all_contacts(self) -> List[Contact]:
        """Get all loaded contacts."""
        if not self._loaded:
            self.load_contacts()
        return list(self._contacts_cache.values())
