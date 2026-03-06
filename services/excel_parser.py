"""Excel parser service for Formación evaluation files."""

import logging
import re
from decimal import Decimal
from io import BytesIO
from typing import List, Optional, Dict, Any, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from models.schemas import (
    Recipient,
    DataEntry,
    DataBlock,
    ColumnInfo,
    ParsedDocument,
)

logger = logging.getLogger(__name__)

# Header row indices (1-based)
HEADER_ROW_1 = 1
HEADER_ROW_2 = 2
HEADER_ROW_3 = 3
HEADER_ROWS = [HEADER_ROW_1, HEADER_ROW_2, HEADER_ROW_3]

# Known label for the Tutor column
TUTOR_LABELS = ["tutor", "tutora"]


class ExcelParser:
    """Parses uploaded Formación evaluation Excel files and splits by Tutor."""

    def __init__(self, mode: str = "tutor"):
        """Initialize the parser.

        Args:
            mode: Split mode identifier (always 'tutor' for this app)
        """
        self.mode = mode

    def parse(self, file_content: bytes, filename: str = "") -> ParsedDocument:
        """Parse Excel file content and return structured document.

        Args:
            file_content: Raw bytes of the uploaded .xlsx file
            filename: Original filename

        Returns:
            ParsedDocument with blocks grouped by Tutor
        """
        wb = load_workbook(BytesIO(file_content), data_only=True)
        ws = wb.active

        try:
            return self._parse_worksheet(ws, filename)
        finally:
            wb.close()

    def get_columns(self, file_content: bytes) -> List[ColumnInfo]:
        """Extract column metadata from Excel file.

        Args:
            file_content: Raw bytes of the uploaded .xlsx file

        Returns:
            List of ColumnInfo with category/subcategory hierarchy
        """
        wb = load_workbook(BytesIO(file_content), data_only=True)
        ws = wb.active
        try:
            merged = self._get_merged_ranges(ws)
            return self._build_column_info(ws, merged)
        finally:
            wb.close()

    def _parse_worksheet(self, ws: Worksheet, filename: str) -> ParsedDocument:
        """Parse the active worksheet into a structured document."""
        merged = self._get_merged_ranges(ws)

        # Detect Tutor column
        tutor_col_idx = self._find_tutor_column(ws)
        tutor_col_letter = get_column_letter(tutor_col_idx) if tutor_col_idx else None

        if not tutor_col_idx:
            raise ValueError(
                "No se encontró la columna 'Tutor' en la fila de encabezados. "
                "Verifique que el archivo contiene una columna etiquetada como 'Tutor' en la fila 3."
            )

        # Find data start row (skip empty separator rows after headers)
        data_start = self._find_data_start(ws)

        # Group rows by Tutor value
        tutor_groups = self._group_by_tutor(ws, tutor_col_idx, data_start)

        # Build column info with categories
        all_columns = self._build_column_info(ws, merged)

        # Default: include all columns
        default_columns = [col.letter for col in all_columns]

        # Determine max column used
        max_col = ws.max_column or 1

        # Build blocks
        blocks: List[DataBlock] = []
        sample_tutor = None

        for tutor_name, row_indices in tutor_groups.items():
            if sample_tutor is None:
                sample_tutor = tutor_name

            recipient = Recipient(codigo=tutor_name, nombre=tutor_name)

            entries = []
            for row_idx in row_indices:
                raw_cells: Dict[str, Any] = {}
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    if value is not None:
                        col_letter = get_column_letter(col_idx)
                        raw_cells[col_letter] = value

                entries.append(DataEntry(
                    raw_cells=raw_cells,
                    source_row=row_idx,
                ))

            block = DataBlock(
                recipient=recipient,
                entries=entries,
                start_row=min(row_indices) if row_indices else 0,
                end_row=max(row_indices) if row_indices else 0,
            )
            blocks.append(block)

        logger.info(
            "Archivo analizado: %d tutores, %d filas de datos",
            len(blocks),
            sum(len(b.entries) for b in blocks),
        )

        return ParsedDocument(
            filename=filename,
            blocks=blocks,
            total_general=Decimal("0"),
            header_rows=HEADER_ROWS,
            all_columns=all_columns,
            default_columns=default_columns,
            tutor_column=tutor_col_letter,
            sample_tutor=sample_tutor,
        )

    def _get_merged_ranges(self, ws: Worksheet) -> Dict[str, Tuple[str, int, int, int, int]]:
        """Get merged cell ranges indexed by top-left cell.

        Returns:
            Dict mapping 'COL_LETTER:ROW' to (value, min_col, min_row, max_col, max_row)
        """
        merged: Dict[str, Tuple[str, int, int, int, int]] = {}
        for merge_range in ws.merged_cells.ranges:
            min_col = merge_range.min_col
            min_row = merge_range.min_row
            max_col = merge_range.max_col
            max_row = merge_range.max_row
            top_left = ws.cell(row=min_row, column=min_col)
            value = top_left.value
            key = f"{get_column_letter(min_col)}:{min_row}"
            merged[key] = (value, min_col, min_row, max_col, max_row)
        return merged

    def _find_tutor_column(self, ws: Worksheet) -> Optional[int]:
        """Auto-detect the Tutor column by searching Row 3 for the label.

        Returns:
            Column index (1-based) or None if not found
        """
        max_col = ws.max_column or 1
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=HEADER_ROW_3, column=col_idx)
            value = cell.value
            if value and isinstance(value, str):
                normalized = value.strip().lower()
                if normalized in TUTOR_LABELS:
                    return col_idx

            # Also check if it's a merged cell with value in Row 2
            cell2 = ws.cell(row=HEADER_ROW_2, column=col_idx)
            if cell2.value and isinstance(cell2.value, str):
                normalized2 = cell2.value.strip().lower()
                if normalized2 in TUTOR_LABELS:
                    return col_idx

        return None

    def _find_data_start(self, ws: Worksheet) -> int:
        """Find the first data row after headers, skipping empty separator rows.

        Returns:
            Row index (1-based) of the first data row
        """
        for row_idx in range(HEADER_ROW_3 + 1, ws.max_row + 1):
            has_data = False
            for col_idx in range(1, min(ws.max_column + 1, 20)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and not (isinstance(cell.value, str) and not cell.value.strip()):
                    has_data = True
                    break
            if has_data:
                return row_idx
        return HEADER_ROW_3 + 2  # Default: row 5

    def _group_by_tutor(
        self, ws: Worksheet, tutor_col: int, data_start: int
    ) -> Dict[str, List[int]]:
        """Group data row indices by their Tutor column value.

        Args:
            ws: Worksheet
            tutor_col: Column index (1-based) of the Tutor column
            data_start: First data row index

        Returns:
            OrderedDict-like dict mapping tutor name -> list of row indices
        """
        from collections import OrderedDict
        groups: Dict[str, List[int]] = OrderedDict()

        for row_idx in range(data_start, ws.max_row + 1):
            # Skip empty rows
            tutor_cell = ws.cell(row=row_idx, column=tutor_col)
            tutor_value = tutor_cell.value

            if tutor_value is None or (isinstance(tutor_value, str) and not tutor_value.strip()):
                continue

            tutor_name = str(tutor_value).strip()

            # Verify the row has some data (not just a Tutor value)
            has_other_data = False
            for col_idx in range(1, min(ws.max_column + 1, 20)):
                if col_idx == tutor_col:
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and not (isinstance(cell.value, str) and not cell.value.strip()):
                    has_other_data = True
                    break

            if has_other_data:
                if tutor_name not in groups:
                    groups[tutor_name] = []
                groups[tutor_name].append(row_idx)

        return groups

    def _build_column_info(
        self, ws: Worksheet, merged: Dict[str, Tuple[str, int, int, int, int]]
    ) -> List[ColumnInfo]:
        """Build column metadata with category (Row 2) and subcategory (Row 3).

        Handles merged cells that span multiple columns in Row 2 (categories)
        and merged cells that span Row 2 and Row 3 (combined category/subcategory).
        """
        max_col = ws.max_column or 1
        columns: List[ColumnInfo] = []

        # Build a map: col_idx -> category from Row 2 (accounting for merged cells)
        category_map: Dict[int, str] = {}
        for key, (value, min_col, min_row, max_col_m, max_row) in merged.items():
            if min_row == HEADER_ROW_2 and value is not None:
                cat_text = str(value).strip()
                for c in range(min_col, max_col_m + 1):
                    category_map[c] = cat_text

        # Also check non-merged cells in Row 2
        for col_idx in range(1, max_col + 1):
            if col_idx not in category_map:
                cell = ws.cell(row=HEADER_ROW_2, column=col_idx)
                if cell.value is not None and not isinstance(cell, MergedCell):
                    if isinstance(cell.value, str) and cell.value.strip():
                        category_map[col_idx] = cell.value.strip()

        # Build column info
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)

            # Get subcategory from Row 3
            header_text = None
            cell3 = ws.cell(row=HEADER_ROW_3, column=col_idx)
            if cell3.value is not None and not isinstance(cell3, MergedCell):
                header_text = str(cell3.value).strip()

            # For cells merged vertically from Row 2 to Row 3, the header IS the category
            if header_text is None:
                # Check if this column is part of a vertical merge (row 2 to row 3)
                for key, (value, min_col, min_row, max_col_m, max_row) in merged.items():
                    if (min_row == HEADER_ROW_2 and max_row >= HEADER_ROW_3 and
                            min_col <= col_idx <= max_col_m):
                        if value is not None:
                            header_text = str(value).strip()
                        break

            category = category_map.get(col_idx)

            # Skip columns with no header at all
            if header_text is None and category is None:
                continue

            # If header equals category (vertically merged), clear category to avoid duplication
            if header_text and category and header_text == category:
                category = None

            columns.append(ColumnInfo(
                letter=col_letter,
                header=header_text,
                category=category,
                has_data=True,
            ))

        return columns
