"""Excel generator service for creating individual files per Tutor."""

import logging
import os
import re
import shutil
import subprocess
import tempfile
import zipfile
from copy import copy
from io import BytesIO
from typing import List, Tuple, Optional, Dict

from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from models.schemas import ParsedDocument, DataBlock

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generates individual Excel files per Tutor preserving original formatting."""

    def generate_files(
        self,
        document: ParsedDocument,
        source_content: bytes,
        selected_columns: List[str],
        original_filename: str = "",
    ) -> List[Tuple[str, bytes]]:
        """Generate individual Excel files for each Tutor.

        Args:
            document: Parsed document with blocks per Tutor
            source_content: Original uploaded file bytes
            selected_columns: Column letters to include in output
            original_filename: Original filename for naming convention

        Returns:
            List of (filename, content_bytes) tuples
        """
        source_wb = load_workbook(BytesIO(source_content), data_only=True)
        # Also load with styles (not data_only) for full formatting
        style_wb = load_workbook(BytesIO(source_content))
        source_ws = source_wb.active
        style_ws = style_wb.active

        files: List[Tuple[str, bytes]] = []

        # Build column mapping: source col letter -> output col index (1-based)
        col_mapping = self._build_column_mapping(selected_columns)

        # Collect merged cell ranges from source
        source_merges = list(style_ws.merged_cells.ranges)

        for block in document.blocks:
            if block.recipient.excluded:
                continue

            try:
                filename, content = self._generate_single_file(
                    source_ws, style_ws, block, col_mapping,
                    source_merges, original_filename,
                )
                files.append((filename, content))
            except Exception as e:
                logger.error("Error generando archivo para tutor %s: %s",
                             block.recipient.nombre, str(e))

        source_wb.close()
        style_wb.close()

        logger.info("Generados %d archivos individuales", len(files))
        return files

    def _generate_single_file(
        self,
        source_ws: Worksheet,
        style_ws: Worksheet,
        block: DataBlock,
        col_mapping: Dict[str, int],
        source_merges: list,
        original_filename: str,
    ) -> Tuple[str, bytes]:
        """Generate a single Excel file for one Tutor."""
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = source_ws.title or "Evaluaciones"

        # Step 1: Copy header rows (1-3) with formatting
        for src_row in range(1, 4):
            dest_row = src_row
            self._copy_row_cells(style_ws, new_ws, src_row, dest_row, col_mapping)
            # Copy row height
            src_height = style_ws.row_dimensions[src_row].height
            if src_height:
                new_ws.row_dimensions[dest_row].height = src_height

        # Step 2: Copy merged cells in header rows (adjusted for column mapping)
        self._copy_header_merges(style_ws, new_ws, col_mapping, source_merges)

        # Step 3: Copy data rows for this Tutor
        dest_row = 4  # Start after headers (no separator row in output)
        for entry in block.entries:
            src_row = entry.source_row
            if src_row <= 0:
                continue
            self._copy_row_cells(style_ws, new_ws, src_row, dest_row, col_mapping)
            # Copy row height
            src_height = style_ws.row_dimensions[src_row].height
            if src_height:
                new_ws.row_dimensions[dest_row].height = src_height
            dest_row += 1

        # Step 4: Copy column widths for selected columns
        self._copy_column_widths(style_ws, new_ws, col_mapping)

        # Generate filename: {Tutor_Name}_{Original_Filename}.xlsx
        tutor_safe = self._sanitize_filename(block.recipient.nombre)
        base_name = original_filename.replace(".xlsx", "").replace(".xls", "")
        base_safe = self._sanitize_filename(base_name) if base_name else "Evaluacion"
        filename = f"{tutor_safe}_{base_safe}.xlsx"

        # Save to bytes
        buffer = BytesIO()
        new_wb.save(buffer)
        buffer.seek(0)
        content = buffer.read()
        new_wb.close()

        return filename, content

    def _copy_row_cells(
        self,
        source_ws: Worksheet,
        dest_ws: Worksheet,
        src_row: int,
        dest_row: int,
        col_mapping: Dict[str, int],
    ) -> None:
        """Copy cells from source row to dest row using column mapping."""
        for src_letter, dest_col_idx in col_mapping.items():
            src_col_idx = column_index_from_string(src_letter)
            source_cell = source_ws.cell(row=src_row, column=src_col_idx)
            dest_cell = dest_ws.cell(row=dest_row, column=dest_col_idx)

            # Skip MergedCell objects (they don't have actual values/styles)
            if isinstance(source_cell, MergedCell):
                continue

            # Copy value
            dest_cell.value = source_cell.value

            # Copy all formatting
            self._copy_cell_style(source_cell, dest_cell)

    def _copy_cell_style(self, source_cell, dest_cell) -> None:
        """Copy all formatting from source cell to destination cell."""
        if source_cell.font:
            dest_cell.font = copy(source_cell.font)
        if source_cell.fill and source_cell.fill.patternType:
            dest_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            dest_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            dest_cell.alignment = copy(source_cell.alignment)
        if source_cell.number_format:
            dest_cell.number_format = source_cell.number_format
        if source_cell.protection:
            dest_cell.protection = copy(source_cell.protection)

    def _copy_header_merges(
        self,
        source_ws: Worksheet,
        dest_ws: Worksheet,
        col_mapping: Dict[str, int],
        source_merges: list,
    ) -> None:
        """Copy merged cell ranges from headers, adjusting for column mapping."""
        selected_col_indices = set()
        for letter in col_mapping:
            selected_col_indices.add(column_index_from_string(letter))

        for merge_range in source_merges:
            min_row = merge_range.min_row
            max_row = merge_range.max_row

            # Only process header merges (rows 1-3)
            if min_row > 3:
                continue

            min_col = merge_range.min_col
            max_col = merge_range.max_col

            # Find which selected columns overlap with this merge range
            mapped_cols = []
            for src_col in range(min_col, max_col + 1):
                src_letter = get_column_letter(src_col)
                if src_letter in col_mapping:
                    mapped_cols.append(col_mapping[src_letter])

            if len(mapped_cols) < 2 and min_row == max_row:
                # Single cell in output — no merge needed for horizontal merges
                continue

            if mapped_cols:
                new_min_col = min(mapped_cols)
                new_max_col = max(mapped_cols)
                # Clamp max_row to 3 (headers only)
                new_max_row = min(max_row, 3)

                if new_min_col < new_max_col or min_row < new_max_row:
                    try:
                        dest_ws.merge_cells(
                            start_row=min_row,
                            start_column=new_min_col,
                            end_row=new_max_row,
                            end_column=new_max_col,
                        )
                    except ValueError:
                        pass  # Already merged or invalid range

    def _copy_column_widths(
        self,
        source_ws: Worksheet,
        dest_ws: Worksheet,
        col_mapping: Dict[str, int],
    ) -> None:
        """Copy column widths from source to destination."""
        for src_letter, dest_col_idx in col_mapping.items():
            src_width = source_ws.column_dimensions[src_letter].width
            dest_letter = get_column_letter(dest_col_idx)
            if src_width:
                dest_ws.column_dimensions[dest_letter].width = src_width
            else:
                dest_ws.column_dimensions[dest_letter].width = 12

    def _build_column_mapping(self, selected_columns: List[str]) -> Dict[str, int]:
        """Build mapping from source column letter to output column index.

        Selected columns are placed sequentially starting at column 1.
        """
        mapping: Dict[str, int] = {}
        # Sort columns by their original position to maintain order
        sorted_cols = sorted(selected_columns, key=lambda l: column_index_from_string(l))
        for idx, letter in enumerate(sorted_cols, start=1):
            mapping[letter] = idx
        return mapping

    def _sanitize_filename(self, name: str) -> str:
        """Create safe filename from a name string."""
        name = name.strip()
        name = re.sub(r'[<>:"/\\|?*]', "_", name)
        name = name.replace(" ", "_")
        name = re.sub(r"_+", "_", name)
        if len(name) > 60:
            name = name[:60]
        return name.strip("_")

    def create_zip_archive(
        self, files: List[Tuple[str, bytes]], zip_filename: str = "evaluaciones.zip"
    ) -> bytes:
        """Create a ZIP archive containing all generated files."""
        buffer = BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for filename, content in files:
                zf.writestr(filename, content)
        buffer.seek(0)
        return buffer.read()

    # ── Screenshot Generation ────────────────────────────────────────

    def generate_screenshots(
        self, files: List[Tuple[str, bytes]]
    ) -> List[Tuple[str, bytes]]:
        """Generate PNG screenshots for each generated Excel file.

        Args:
            files: List of (filename, content_bytes) from generate_files

        Returns:
            List of (png_filename, png_bytes) tuples
        """
        screenshots: List[Tuple[str, bytes]] = []
        for filename, content in files:
            try:
                png_name = filename.replace(".xlsx", ".png")
                png_content = self._render_excel_to_png(content)
                screenshots.append((png_name, png_content))
            except Exception as e:
                logger.error(
                    "Error generando captura para %s: %s", filename, str(e)
                )
        logger.info("Generadas %d capturas de pantalla", len(screenshots))
        return screenshots

    def _render_excel_to_png(self, file_content: bytes) -> bytes:
        """Render Excel file content as a PNG image.

        Uses LibreOffice in headless mode to convert Excel → PDF,
        then PyMuPDF (fitz) to render the PDF page(s) as a high-quality PNG.
        This preserves all original formatting: colors, borders, merged
        cells, fonts, and displays computed formula values.

        Before conversion, the sheet's page setup is adjusted to landscape
        orientation with fit-to-width scaling so that wide spreadsheets are
        not truncated across multiple PDF pages.
        """
        import fitz  # PyMuPDF
        from PIL import Image

        soffice = self._find_libreoffice()
        if not soffice:
            raise RuntimeError(
                "LibreOffice no encontrado. Instálelo para generar capturas. "
                "https://www.libreoffice.org/download/"
            )

        tmpdir = tempfile.mkdtemp(prefix="formacion_screenshot_")
        try:
            # 1. Pre-process: set page layout to landscape + fit-to-width
            prepared_content = self._prepare_xlsx_for_export(file_content)

            xlsx_path = os.path.join(tmpdir, "input.xlsx")
            with open(xlsx_path, "wb") as f:
                f.write(prepared_content)

            # 2. Convert XLSX → PDF via LibreOffice headless
            cmd = [
                soffice,
                "--headless",
                "--calc",
                "--convert-to", "pdf",
                "--outdir", tmpdir,
                xlsx_path,
            ]
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=60,
            )
            if result.returncode != 0:
                logger.error(
                    "LibreOffice conversion failed: %s", result.stderr
                )
                raise RuntimeError(
                    f"LibreOffice conversion failed: {result.stderr}"
                )

            pdf_path = os.path.join(tmpdir, "input.pdf")
            if not os.path.exists(pdf_path):
                raise RuntimeError(
                    "PDF output not found after LibreOffice conversion"
                )

            # 3. Render PDF pages as PNG via PyMuPDF
            doc = fitz.open(pdf_path)
            zoom = 4.0  # 400 DPI equivalent for maximum resolution output
            mat = fitz.Matrix(zoom, zoom)

            if doc.page_count == 1:
                # Single page — render directly
                page = doc.load_page(0)
                pix = page.get_pixmap(matrix=mat, alpha=False)
                png_bytes = pix.tobytes("png")
            else:
                # Multiple pages — stitch vertically into one image
                page_images = []
                for page_num in range(doc.page_count):
                    page = doc.load_page(page_num)
                    pix = page.get_pixmap(matrix=mat, alpha=False)
                    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    page_images.append(img)

                total_width = max(img.width for img in page_images)
                total_height = sum(img.height for img in page_images)
                combined = Image.new("RGB", (total_width, total_height), "white")

                y_offset = 0
                for img in page_images:
                    combined.paste(img, (0, y_offset))
                    y_offset += img.height

                buffer = BytesIO()
                combined.save(buffer, format="PNG", optimize=True)
                buffer.seek(0)
                png_bytes = buffer.read()

            doc.close()

            # 4. Auto-crop whitespace around the content
            png_bytes = self._crop_whitespace(png_bytes)

            return png_bytes

        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)

    @staticmethod
    def _crop_whitespace(png_bytes: bytes, padding: int = 10) -> bytes:
        """Remove excess whitespace around the content in a PNG image.

        Detects the bounding box of non-white content and crops with
        a small padding to produce a tight, clean screenshot.
        """
        from PIL import Image, ImageChops

        img = Image.open(BytesIO(png_bytes)).convert("RGB")

        # Create a white background and diff to find content
        bg = Image.new("RGB", img.size, (255, 255, 255))
        diff = ImageChops.difference(img, bg)
        bbox = diff.getbbox()

        if bbox:
            # Add padding around the content
            left = max(0, bbox[0] - padding)
            top = max(0, bbox[1] - padding)
            right = min(img.width, bbox[2] + padding)
            bottom = min(img.height, bbox[3] + padding)
            img = img.crop((left, top, right, bottom))

        buffer = BytesIO()
        img.save(buffer, format="PNG", optimize=True)
        buffer.seek(0)
        return buffer.read()

    @staticmethod
    def _prepare_xlsx_for_export(file_content: bytes) -> bytes:
        """Adjust the XLSX page setup so LibreOffice exports all content.

        Sets landscape orientation, removes margins, and configures
        fit-to-width so wide spreadsheets are not truncated.
        """
        from openpyxl.worksheet.page import PageMargins

        wb = load_workbook(BytesIO(file_content))
        ws = wb.active

        # Landscape orientation
        ws.page_setup.orientation = "landscape"

        # Fit all columns onto one page width (height can span pages)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        # Minimal margins
        ws.page_margins = PageMargins(
            left=0.2, right=0.2, top=0.2, bottom=0.2,
            header=0.1, footer=0.1
        )

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        wb.close()
        return buf.read()

    @staticmethod
    def _find_libreoffice() -> Optional[str]:
        """Locate the LibreOffice (soffice) executable.

        Searches common installation paths on Windows, Linux, and macOS.
        Returns the path to the executable, or None if not found.
        """
        # 1. Check if soffice / libreoffice is on PATH
        for name in ("soffice", "libreoffice"):
            path = shutil.which(name)
            if path:
                return path

        # 2. Check common installation directories
        candidates = [
            # Windows
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            # Linux
            "/usr/bin/libreoffice",
            "/usr/bin/soffice",
            "/usr/lib/libreoffice/program/soffice",
            # macOS
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        ]
        for candidate in candidates:
            if os.path.isfile(candidate):
                return candidate

        return None
